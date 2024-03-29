# -*- coding: utf-8 -*-
# @Author: chunyang.xu
# @Date:   2022-05-20 18:59:04
# @Last Modified by:   longfengpili
# @Last Modified time: 2023-06-07 15:10:10

import sys
import openpyxl

import logging
elogger = logging.getLogger(__name__)


class ReadDataFromExcel(object):

    def __init__(self, filepath: str, data_only: bool = True):
        """[summary]
        
        [description]
        
        Arguments:
            filepath {str} -- [文件地址]
        
        Keyword Arguments:
            data_only {bool} -- [是否读取公式结果] (default: {True})
        """
        
        self.filepath = filepath
        self.data_only = data_only

    def open_excel(self):
        try:
            book = openpyxl.load_workbook(self.filepath, data_only=self.data_only)
            return book
        except Exception as e:
            elogger.error(f"{e}")
            sys.exit(0)

    def get_sheets(self):
        book = self.open_excel()
        sheetnames = book.sheetnames
        sheets = dict([(sheetname, book[sheetname]) for sheetname in sheetnames])
        return sheets

    def open_sheet(self, sheetname: str):
        book = self.open_excel()
        sheet = book[sheetname]
        return sheet

    def get_sheetvalues_by_rows(self, sheetname: str):
        """[summary]
        
        [以行的模式获取sheetvalues]
        
        Arguments:
            sheetname {str} -- [sheetname]
        
        Returns:
            [list] -- [row value list]
        """

        sheet = self.open_sheet(sheetname)
        srow_values = [row for row in sheet.iter_rows(values_only=True)]
        return srow_values

    def get_sheetvalues_by_cols(self, sheetname: str):
        """[summary]
        
        [以列的模式获取sheetvalues]
        
        Arguments:
            sheetname {str} -- [sheetname]
        
        Keyword Arguments:
            col_num {int} -- [获取的列数] (default: {None})
        
        Returns:
            [list] -- [col value list]
        """

        sheet = self.open_sheet(sheetname)
        scol_values = [col for col in sheet.iter_cols(values_only=True)]
        return scol_values

    def get_sheet_values_by_header(self, sheetname, headers, header_row=0):
        """[summary]

        [根据列名选择]

        Arguments:
            sheetname {[str]} -- [表格名]
            headers {[list]} -- [要选择的列名]

        Keyword Arguments:
            header_row {number} -- [header所在的row] (default: {0})

        Returns:
            [type] -- [description]
        """
        
        srow_values = self.get_sheetvalues_by_rows(sheetname)
        srow_headers = srow_values[header_row]
        headers_index = [srow_headers.index(header) for header in headers]

        scol_values = self.get_sheetvalues_by_cols(sheetname)
        headers_values = [scol_values[hid] for hid in headers_index]

        headers_values = list(zip(*headers_values))
        return headers_values


if __name__ == '__main__':
    filepath = './test.xlsx'
    rexcel = ReadDataFromExcel(filepath)
    srow_values = rexcel.get_sheetvalues_by_rows('first')
    print(srow_values)

    scol_values = rexcel.get_sheetvalues_by_rows('second')
    print(scol_values)